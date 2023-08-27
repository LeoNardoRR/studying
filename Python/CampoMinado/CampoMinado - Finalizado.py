#EXERCÍCIO 04 - PROGRAMAÇÃO 2
#Alunos: VICTOR HENRIQUE DE ALVARENGA e LEONARDO RODRIGUES RIBEIRO
#Feito pelo programa VSCODE

def IniciarJogo():
      #print('---- DEF IniciarJogo() iniciada ----')
      print ('Olá jogador(a), bem vindo!')
      nome = (input('Qual seria seu nome?\nNome: '))
      print ('Vamos jogar então: ', nome)
def CriadorDeCampos():
      #print('---- DEF CriadorDeCampos() iniciada ----')
      from openpyxl import Workbook
      arquivo_excel = Workbook()
      planilha = arquivo_excel.active
      planilha.title ='CAMPO_SECRETO'
      planilha1 = arquivo_excel.create_sheet("CAMPO_TELA")        
      arquivo_excel.save('Campo_minado.xlsx') 
      arquivo_excel.close
def DefinicaoDeCampo():
      #print('--- DEF DefinicaoDeCampo() iniciada ----')
      while True:
            try: 
                  largura = int(input('Digite a quantidade de colunas do campo minado: '))
                  altura = int(input('Digite a quantidade de linhas do campo minado: '))
                  
                  if altura < 3 or largura < 3:
                        print('As colunas e linhas devem ser maiores do que 3')
                        print()
                        continue
                  #print('--- DEF DefinicaoDeCampo() finalizada ----')
                  Nivel(largura, altura)
                  
            except:
                  print('Digite um valor inteiro valido!')
                  print()
                  continue
            break
def Nivel(largura, altura):
      #print('--- DEF Nivel() iniciado ---')
      Nivel = 0
      while True:
            print()
            print('Digite 1 para o nível easy!')
            print('Digite 2 para o nível medium!')
            print('Digite 3 para o nível hard!')
            print('Escolha o nível de Dificuldade')

            Nivel = int(input('Digite 1, 2 ou 3: '))
            
            if Nivel == 1:
                  minas = 0.15 * (largura * altura) 
                  
            elif Nivel == 2:
                  minas = 0.30 * (largura * altura)
                   
            elif Nivel == 3:
                  minas = 0.50 * (largura * altura)
                  
            else:
                  print('digite 1, 2 ou 3, para a escolha do Nivel')
                  continue
            minas = int(minas)
            #print('--- DEF Nivel() finalizada ---')
            MinasCampo(largura, altura, minas)
            break
def PosicaoAleatoriaMina(aleatorio):
      #print('--- DEF PosicaoAleatoriaMina() iniciada ---')
      import random 
      sorteado = random.randint(1, aleatorio)
      #print('--- DEF PosicaoAleatoriaMina() finalizada ---')
      return sorteado
def MinasCampo(largura, altura, minas):
      #print('--- DEF MinasCampo() iniciada ---')
      from openpyxl import load_workbook
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoSecreto = arquivo_excel['CAMPO_SECRETO']
      cont = 1
      while cont <= minas:
            linha = PosicaoAleatoriaMina(largura)
            coluna = PosicaoAleatoriaMina(altura)
            if campoSecreto.cell(row = linha, column = coluna).value != '*':
                  campoSecreto.cell(row = linha, column = coluna).value = '*'
                  cont += 1
      arquivo_excel.save('Campo_minado.xlsx')
      arquivo_excel.close
      #print('--- DEF MinasCampo() finalizada ---')
      ContarMinasProximas(largura, altura,minas)
def ContarMinasProximas(largura, altura,minas):
      #print('--- DEF ContarMinasProximas() iniciada ---')
      from openpyxl import load_workbook
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoSecreto = arquivo_excel['CAMPO_SECRETO']
      for i in range(1, altura+1):
            for j in range(1, largura+1):
                  cont = 0
                  for a in range(i-1, i + 2):
                        for l in range(j - 1, j + 2):
                              if a >=1:
                                    if l >= 1:
                                          if campoSecreto.cell(row = a, column = l).value == '*':
                                                cont += 1
                  if  campoSecreto.cell(row = i, column = j).value != '*':
                        campoSecreto.cell(row = i, column = j).value = str(cont)
      arquivo_excel.save('Campo_minado.xlsx')
      arquivo_excel.close
      #print('--- DEF ContarMinasProximas() finalizada ---')
      preencherCampoTela(largura, altura,minas)
def preencherCampoTela(largura, altura,minas):
      #print('--- DEF preencherCampoTela() iniciada ---')
      from openpyxl import load_workbook
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoTela = arquivo_excel['CAMPO_TELA']
      abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','w','X','y','Z']
      for i in range(1, altura+1):
            for j in range(1, largura+1):
                  campoTela.cell(row = i, column = j).value = '-'
                  if j == largura:
                        campoTela.cell(row = i, column = j+1).value = i
                  if i == altura:
                        campoTela.cell(row = i+1, column = j).value = abc[j-1]
      arquivo_excel.save('Campo_minado.xlsx')
      arquivo_excel.close
      #print('--- DEF preencherCampoTela() finalizada ---')
      InserirJogada(largura, altura,minas)
def verJogoBomba(largura, altura):
      #print('--- DEF verJogo() iniciada ---')
      from openpyxl import load_workbook
      
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoTela = arquivo_excel['CAMPO_TELA']
      campoSecreto = arquivo_excel['CAMPO_SECRETO']
      for i in range(1, campoTela.max_row + 1):
            verJogo = []
            for j in range(1, campoTela.max_column + 1):
                
                if campoSecreto.cell(row = altura, column = largura) == campoSecreto.cell(row = i, column = j) and campoSecreto.cell(row = i, column = j).value == '*':
                        verJogo.append('*')
                elif campoTela.cell(row = i, column = j).value != None:
                    verJogo.append(campoTela.cell(row = i, column = j).value)

            print(verJogo)     
def verJogo(largura, altura):
      #print('--- DEF verJogo() iniciada ---')
      from openpyxl import load_workbook
      
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoTela = arquivo_excel['CAMPO_TELA']
      campoSecreto = arquivo_excel['CAMPO_SECRETO']
      for i in range(1, campoTela.max_row + 1):
            verJogo = []
            for j in range(1, campoTela.max_column + 1):
  
                if campoTela.cell(row = i, column = j).value != None:
                    verJogo.append(campoTela.cell(row = i, column = j).value)

            print(verJogo)  
def InserirJogada(largura, altura,minas):
      #print('--- DEF InserirJogada() iniciada ---')
      contagemParaVencer = 0
      while True:
            # mostra a tela para o jogador
            verJogo(largura, altura)
            abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','w','X','y','Z',
            'A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1','w1','X1','y1','Z1']
            
            try:
                  converterJogadaLargura = str(input('Digite a coluna escolhida ')).upper()
                  jogadaLargura = abc.index(converterJogadaLargura) + 1
                  jogadaAltura = int(input('Digite a linha escolhida '))            
                  if jogadaLargura < 1 or jogadaLargura > 26 or jogadaAltura < 1 or jogadaAltura > 26:
                        print('Digite uma jogada valida')
                        print()
                        continue
                  g = conferirJogada(largura, altura, jogadaLargura, jogadaAltura)
                  Vitoria(minas)

                  if g == True:
                    break
                  
            except:
                  print('Digite uma jogada valida')
                  print()
                  continue
def conferirJogada(largura, altura, jogadaLargura, jogadaAltura):
      
      #print('--- DEF conferirJogada() iniciada ---')
      from openpyxl import load_workbook
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoTela = arquivo_excel['CAMPO_TELA']
      campoSecreto = arquivo_excel['CAMPO_SECRETO']
      valorCampoSecreto = campoSecreto.cell(row = jogadaAltura, column = jogadaLargura).value
      if valorCampoSecreto == '*':
            print('Game Over')
            verJogoBomba(jogadaLargura, jogadaAltura)
            return True
            
      elif campoTela.cell(row = jogadaAltura, column = jogadaLargura).value != '-':
            print('Jogada repetida, escolha outra posição')
      elif campoTela.cell(row = jogadaAltura, column = jogadaLargura).value == '-':
            if valorCampoSecreto == '0':
                  campoTela.cell(row = jogadaAltura, column = jogadaLargura).value = valorCampoSecreto
                  arquivo_excel.save('Campo_minado.xlsx')
                  ValorZero(True)
            else:
                  campoTela.cell(row = jogadaAltura, column = jogadaLargura).value =  valorCampoSecreto
                  arquivo_excel.save('Campo_minado.xlsx')
      else:
            print('Digite uma jogada valida')
            #print('--- DEF vconferirJogada() finalizada ---')
            InserirJogada(largura, altura)
def ValorZero(sair):
      
      from openpyxl import load_workbook
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoTela = arquivo_excel['CAMPO_TELA']
      campoSecreto = arquivo_excel['CAMPO_SECRETO']     
      while True:
            for i in range(1, campoTela.max_row):
                  for j in range(1,  campoTela.max_column):
                        for a in range(i-1, i + 2):
                              for l in range(j - 1, j + 2):
                                    if a >=1 and a <= campoTela.max_row-1:
                                          if l >= 1 and l <= campoTela.max_column-1:
                                                if campoTela.cell(row = i, column = j).value == '0':
                                                      if campoSecreto.cell(row = a, column = l).value != '*':
                                                            campoTela.cell(row = a, column = l).value= campoSecreto.cell(row = a, column = l).value                                                                                                            
                                                            arquivo_excel.save('Campo_minado.xlsx')

            arquivo_excel.close
            if sair == True:
                  sair = VerificarZero(sair)
            else:
                  break
def VerificarZero(sair):
      from openpyxl import load_workbook
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoTela = arquivo_excel['CAMPO_TELA']
      campoSecreto = arquivo_excel['CAMPO_SECRETO']
      for i in range(1, campoTela.max_row):
            for j in range(1,  campoTela.max_column):
                  for a in range(i-1, i + 2):
                        for l in range(j - 1, j + 2):
                              if a >=1 and a <= campoTela.max_row-1:
                                    if l >= 1 and l <= campoTela.max_column-1:
                                          if campoTela.cell(row = i, column = j).value == '0':
                                                if campoTela.cell(row = a, column = l).value == '-':
                                                      if campoSecreto.cell(row = a, column = l).value != '*':
                                                            return True
      
      return False
def Vitoria(minas):

      from openpyxl import load_workbook
      arquivo_excel = load_workbook('Campo_minado.xlsx', read_only = False)
      campoTela = arquivo_excel['CAMPO_TELA']
      cont = 0
      for i in range(1, campoTela.max_row):
            for j in range(1,  campoTela.max_column):
                  if campoTela.cell(row = i, column = j).value == '-':
                        cont += 1
      if (cont-minas) == 0:
            print('Você ganhou !!!!!!')
            
            CriadorDeCampos()
            DefinicaoDeCampo()

while True:
    IniciarJogo()
    CriadorDeCampos()
    DefinicaoDeCampo()

    op = input('Deseja continuar? (S or N) ').upper()
    
    if op == 'N':
        break
